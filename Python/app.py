# First Programme
print('Akash Deshbhratar')
print('o----')
print('  ||||')
print('*' * 10)

# Second Programme

price = 10  # number
price = 20  # number
rating = 4.9  # floating point number
name = 'Akash'  # string
is_published = False  # boolean
print(is_published)
print(rating)
print(price)

full_name = 'John Doe'
age = 20
is_new = True

# Third Programme

# name = input('What is your name?')
# print('Hi ' + name)
#
# color = input('What is your color?')
#
# print('Hi ' + name + ', Your favourite color is '+ color)

# Fourth Programme
#
# birth_year = input('Birth Year: ')
# print(type(birth_year))
# age = 2019 - int(birth_year) # int(), float(), bool()
# print(age)
# print(type(age))
#
# weight_lbs = input('Weight (lbs):')
# weight_kg = float(weight_lbs) * 0.45
# print(weight_kg)

# Fifth Programme
course = "Python's Course for Beginners"
course1 = 'Python Course for "Beginners"'

multiline_string = '''
    Hi Akash,
    You need to be more productive.
    
    Thank You
'''

print(course)
print(course1)
print(multiline_string)
print(course[0])
print(course[-2])
print(course[0:4])
print(course[1:])
print(course[:5])  # Assums start index as 0
print(course[:])
last_name = "Deshbhratar"
print(last_name[1:-1])

f_name = 'John'
l_name = 'Doe'
message = f_name + ' [' + l_name + '] is a coder'
msg = f'{f_name} [{l_name}] is a coder'
print(message)
print(msg)

# Sixth Programme
course = 'Python for Beginners'
print(len(course))
print(course.upper())
print(course.lower())
print(course.find('o'))
print(course.replace('Beginners', 'Absolute Beginners'))
print('Python' in course)
print(course.title())

# Seventh Programme
print(f'Add = {10 + 2}')
print(f'Sub = {10 - 2}')
print(f'Mul = {10 * 2}')
print(f'div = {10 / 3}')
print(f'Div = {10 // 2}')
print(f'Power = {10 ** 2}')

print(f'Operrator Precendence {10 + 3 * 2 ** 2}')
print((2 + 3) * 10 - 3)
print(round(2.99))
print(abs(-2.9))

import math

print(math.ceil(2.9))
print(math.floor(2.9))

# Eighth Programme
is_hot = False
is_cold = False
if is_hot:
    print("It's a hot day.")
    print("Drink plenty of water")
elif is_cold:
    print("It's a cold day")
    print("Wear warm cloth")
else:
    print("It's a lovely day")
print("Enjoy your day")

price = 1000000
has_good_credit = True
if has_good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.2 * price
print(f"Down Payment: ${price}")

# Ninth Programme
has_high_income = False
has_good_credit = True
has_criminal_record = False
if has_good_credit and has_high_income:
    print("Eligible for loan!!!")

if has_good_credit or has_high_income:
    print("Eligible for loan!!!")

if has_good_credit and not has_criminal_record:
    print("Eligible for loan!!!!")
    # AND: both
    # OR : at leat one

# Tenth programme
temprature = 35
if temprature > 30:
    print("It's a hot day")
else:
    print("It's not a hot day")

if temprature < 30:
    print("It's a hot day")
else:
    print("It's not a hot day")

if temprature >= 30:
    print("It's a hot day")
else:
    print("It's not a hot day")

if temprature <= 30:
    print("It's a hot day")
else:
    print("It's not a hot day")

if temprature == 30:
    print("It's a hot day")
else:
    print("It's not a hot day")

_name = "Jddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddddd"
print(len(_name))
if len(_name) < 3:
    print("Name must be at least 3 charecters")
elif len(_name) > 50:
    print("Name must be a maximum of 50 charecters")
else:
    print("Name looks good")

# weight = input("Enter your weight")
# unit = input("Enter unit Kg(K or k)/Lbs(L or l)")
#
# if unit.upper() == 'L':
#     print(f"your weight is: { float(weight) * 0.45 } Kg")
# elif unit.upper() in 'K':
#     print(f"your weight is: { float(weight) / 0.45} Lbs")
# else:
#     print("Wrong choice")

# Eleventh Programme
i = 1
while i <= 5:
    print("*" * i)
    i = i + 1
print("Done")

# secret_number = 9
# guess_count = 0
# guess_limit = 3
# while guess_count < guess_limit:
#     guess = int(input('Guess: '))
#     guess_count += 1
#     if guess == secret_number:
#         print('You won!')
#         break
# else:
#     print("You lose!!!")


# command = ""
# started = False
# while command != 'quit':
#     command = input("> ").lower()
#     if command == "start":
#         if started:
#             print("Car is already started")
#         else:
#             started = True
#             print("Car started!")
#     elif command == "stop":
#         if not started:
#             print("Car is already stopped!")
#         else:
#             started = False
#             print("Car Stopped!")
#     elif command == "help":
#         print("""
# start - to start the car
# stop  - to stop the car
# quit  - to quit
#             """)
#     elif command == "quit":
#         break
#     else:
#         print("Sorry, I didn't understand")

# 12th Prog
for item in ['Mosh', 'John', 'Sarha']:
    print(item)

for item in [1, 2, 3, 4, 5, 6]:
    print(item)

print("\n")

for item in range(10):
    print(item)

print("\n")

for item in range(5, 10):
    print(item)

print("\n")

# for step 2
for item in range(5, 10, 2):
    print(item)

prices = [10, 20, 30]
total = 0
for price in prices:
    total += price

print(f"Total Price: {total}")

# 13th Prog
for x in range(4):
    for y in range(3):
        print(f"({x}, {y})")

numbers = [5, 2, 5, 2, 2]
for x in numbers:
    opt = ""
    for y in range(x):
        opt += "X"
    print(opt)

print("\n")

# 15th Prog

names = ['Mosh', 'John', 'Sarha', 'Bob', 'Merry']
print(names[2:])
print(names[2:4])
print(names[:])

# Exec
numberss = [311, 6, 7, 8, 93, 2]
max = numberss[0]

for item in numberss:
    if item > max:
        max = item
print(max)

# 16th Prog

matrix = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
for row in matrix:
    for item in row:
        print(item)

numbers = [5, 2, 1, 2, 7, 4]
numbers.append(10)
print(numbers)
numbers.insert(0, 10)
print(numbers)
numbers.remove(5)
print(numbers)
numbers.pop()
print(numbers)
print(20 in numbers)
print(numbers.count(5))
print(numbers.sort())
print(numbers)
numbers.reverse()
print(numbers)
numbers2 = numbers.copy()
print(numbers2)
numbers.clear()
print(numbers)

numbers = [2, 3, 4, 4, 5, 5, 7, 8]
uniques = []
for num in numbers:
    if num not in uniques:
        uniques.append(num)
print(uniques)

# 17th prog
numbers = (1, 2, 3)
print(numbers[0])

# 18th Prog
coordinates = (1, 2, 3)  # tupple example
x = coordinates[0]
y = coordinates[1]
z = coordinates[2]
print(x * y * z)
# OR
x, y, z = coordinates
print(x * y * z)

# 19th Prog (Dictoionary)

customer = {"name": "John Smith", "age": 30, "is_verified": True}
print(customer.get('name'))
print(customer['name'])
print(customer.get('birth_date', 'Jan 1 1990'))
customer['name'] = "Jack Smith"
print(customer['name'])
customer['birth_date'] = "Aug 29 1993"
print(customer['birth_date'])


# phone  = input("Phone: ")
# digits_mapping = {
#     "1":"One",
#     "2": "Two",
#     "3": "Three",
#     "4": "Four"
# }
#
# output = ""
# for ch in phone:
#     output += digits_mapping.get(ch, "!") + " "

# print(output)

# 20th Prog()
# message = input(">")
# words = message.split(" ")
# emojis = {
#     ":)": "😀️👍👍👍",
#     ":(": "😕️"
# }
#
# output = ""
# for word in words:
#     output += emojis.get(word, word) + " "
# print(output)


# 21st Prog(Functions)


def greet_user():  # without params
    print("Hi There!")
    print("Welcome aboard")


greet_user()


def greet_real_user(name):  # With params
    print(f"Hi {name}!")
    print("Welcome aboard")


greet_real_user("Akash")  # With agrument


def greet_full_name_user(f_name, l_name):  # With params
    print(f"Hi {f_name} {l_name}!")
    print("Welcome aboard")


greet_full_name_user(l_name="Deshbhratar", f_name="Akash")  # With agrument


def square(num):
    return num * num


print(square(4))


def square_none(num):
    print(num * num)


print(square_none(4))


def emoji_converter(msg):
    words = msg.split(" ")
    emojis = {
        ":)": "😀️👍👍👍",
        ":(": "😕️"
    }

    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


# print(emoji_converter(input(">")))


# 22nd Prog(Error Handling)

# try:
#     age = int(input('Age: '))
#     print(age)
# except ValueError:
#     print('Error Encounter: Due to invalid Value')
#
# try:
#     age = int(input('Age: '))
#     income = 20000
#     risk = income/age
#     print(age)
# except ValueError:
#     print('Invalid Value')
# except ZeroDivisionError:
#     print('Age can not be 0.')


# 23rd Prog(Classes)
# Numbers, Strings, Booleans are classes
class Point:
    def move(self):
        print('move')

    def draw(self):
        print('draw')


point1 = Point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()


class OPoint:
    def __init__(self, x, y):  # constructor
        self.x = x
        self.y = y

    def move(self):
        print('move')

    def draw(self):
        print('draw')


opoint1 = OPoint(10, 20)
opoint1.x = 11
print(opoint1.x)


class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f'Hi, I am {self.name}, The king of North')


john = Person("Joh Snow")
john.talk()


# Inheritance example

class Mammal:
    def walk(self):
        print('Walk')


class Dog(Mammal):
    def bark(self):
        print('Bhau Bhau')


class Cat(Mammal):
    pass


dog1 = Dog()
dog1.walk()
dog1.bark()

# 24th Prog(Modules->each file call as module)

import converters  # import entire object
from converters import lbs_to_kg  # import specific function or class

print(lbs_to_kg(200))
print(converters.kg_to_lbs(70))
print(converters.lbs_to_kg(10))

import maximum_number
from maximum_number import find_maximum_number

print(find_maximum_number([1, 2, 6, 3, 9]))
print(maximum_number.find_maximum_number([10, 32, 9, 78, 888]))

import ecommerce.shipping
ecommerce.shipping.calc_shipping()

from ecommerce import shipping
shipping.calc_shipping()

from ecommerce.shipping import calc_shipping
calc_shipping()

import random
for i in range(3):
    print(random.random())          # between 0 and 1
    print(random.randint(10, 20))    # between 10 and 20

members = ['John', 'Mary', 'Bob', 'Mosh']
leader = random.choice(members)
print(leader)


class Dice:

    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        # return (first, second)    #tupple example
        return first, second        #tupple example


dice = Dice()
print(dice.roll())



from pathlib import Path

# Absolute Path (from compouter directory)
# Relative Path (current directory)

path = Path("ecommerce")
print(path.exists())

for ff in path.glob('*'):       #Find all files and folders
    print(ff)

for ff in path.glob('*.*'):     #Find all files and folders
    print(ff)

for ff in path.glob('*.py'):    #Find all python files
    print(ff)

path = Path("ecommerce1")
print(path.exists())
print(path.mkdir())     #create directory 'ecommerce1'
print(path.rmdir())     #remove directory 'ecommerce1'




import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet =  wb['Sheet1']
cell = sheet['a1']
sheet.cell(1, 1)
# print(cell.value)
# print(sheet.max_row)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    correct_value = float(cell.value) * 0.9
    correct_price_cell = sheet.cell(row, 4)
    correct_price_cell.value = correct_value

# chart_values = Reference(sheet,
#           min_row=2,
#           max_row=sheet.max_row,
#           min_col=4,
#           max_col=4)

# chart = BarChart()
# chart.add_data(chart_values)
# sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')


























